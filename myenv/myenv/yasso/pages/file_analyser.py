"""
FlowNest ERP - File Analysis Utilities
Reads file content and returns structured data for charts and reports.
"""
import csv
import os
import io
import base64
from django.conf import settings
import google.generativeai as genai
from PIL import Image
from groq import Groq

if settings.GOOGLE_API_KEY:
    genai.configure(api_key=settings.GOOGLE_API_KEY)


# ─── helpers ────────────────────────────────────────────────────────────────

def _is_numeric(val):
    try:
        float(str(val).replace(',', '').strip())
        return True
    except (ValueError, TypeError):
        return False


def _col_numeric_data(rows, col_idx):
    """Return list of floats for one column, skipping non-numeric cells."""
    out = []
    for r in rows:
        if col_idx < len(r):
            v = str(r[col_idx]).replace(',', '').strip()
            try:
                out.append(float(v))
            except ValueError:
                out.append(None)
        else:
            out.append(None)
    return out


# ─── CSV ─────────────────────────────────────────────────────────────────────

def analyse_csv(filepath):
    try:
        with open(filepath, newline='', encoding='utf-8', errors='replace') as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            all_rows = []
            for idx, r in enumerate(reader):
                if idx < 1000: # Limit for analysis speed
                    all_rows.append(r)
                else:
                    break
            rows = all_rows
            total_rows = len(rows)
    except Exception as e:
        return {'type': 'csv', 'error': str(e)}

    # Column stats
    col_stats = []
    for i, col in enumerate(headers):
        filled = sum(1 for r in rows if i < len(r) and str(r[i]).strip())
        numeric_vals = [float(str(r[i]).replace(',', '')) for r in rows
                        if i < len(r) and _is_numeric(r[i])]
        stat = {
            'name': col,
            'filled': filled,
            'pct': round(filled / max(total_rows, 1) * 100),
            'is_numeric': bool(numeric_vals),
            'min': round(min(numeric_vals), 2) if numeric_vals else None,
            'max': round(max(numeric_vals), 2) if numeric_vals else None,
            'avg': round(sum(numeric_vals) / len(numeric_vals), 2) if numeric_vals else None,
            'total': round(sum(numeric_vals), 2) if numeric_vals else None,
        }
        col_stats.append(stat)

    # Build chart datasets from numeric columns (first label col + up to 3 numeric cols)
    label_col = None
    numeric_cols = []
    for i, col in enumerate(col_stats):
        if not col['is_numeric'] and label_col is None:
            label_col = i
        elif col['is_numeric']:
            numeric_cols.append(i)

    chart_labels = [str(r[label_col]) if label_col is not None and label_col < len(r) else f"صف {j+1}"
                    for j, r in enumerate(rows[:50])]   # max 50 points

    chart_datasets = []
    palette = ['#6366f1', '#10b981', '#f59e0b', '#ec4899', '#3b82f6']
    for k, col_i in enumerate(numeric_cols[:4]):   # max 4 numeric cols
        data = _col_numeric_data(rows[:50], col_i)
        chart_datasets.append({
            'label': headers[col_i] if col_i < len(headers) else f'عمود {col_i}',
            'data': data,
            'color': palette[k % len(palette)],
        })

    return {
        'type': 'csv',
        'total_rows': total_rows,
        'total_cols': len(headers),
        'headers': headers,
        'col_stats': col_stats,
        'sample': rows[:5],
        'chart_labels': chart_labels,
        'chart_datasets': chart_datasets,
        'has_charts': bool(chart_datasets),
        'error': None,
    }


# ─── EXCEL ───────────────────────────────────────────────────────────────────

def analyse_excel(filepath):
    try:
        import openpyxl
    except ImportError:
        return {'type': 'excel', 'error': 'pip install openpyxl'}

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames
        result_sheets = []

        palette = ['#6366f1', '#10b981', '#f59e0b', '#ec4899', '#3b82f6']

        for sheet_name in sheets[:2]: # Only scan first 2 sheets for speed
            ws = wb[sheet_name]
            # Limit row scan to first 500 rows for analysis speed
            all_rows = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx > 500: break
                all_rows.append(row)
                
            if not all_rows:
                continue

            headers = [str(h) if h is not None else '' for h in all_rows[0]]
            data_rows = all_rows[1:]
            total_rows = len(data_rows)

            col_stats = []
            for i, col in enumerate(headers):
                vals = [r[i] for r in data_rows if i < len(r) and r[i] is not None]
                num_vals = [float(str(v).replace(',', '')) for v in vals if _is_numeric(v)]
                filled = len(vals)
                col_stats.append({
                    'name': col,
                    'filled': filled,
                    'pct': round(filled / max(total_rows, 1) * 100),
                    'is_numeric': bool(num_vals),
                    'min': round(min(num_vals), 2) if num_vals else None,
                    'max': round(max(num_vals), 2) if num_vals else None,
                    'avg': round(sum(num_vals) / len(num_vals), 2) if num_vals else None,
                    'total': round(sum(num_vals), 2) if num_vals else None,
                })

            # chart data
            label_col = next((i for i, c in enumerate(col_stats) if not c['is_numeric']), None)
            numeric_col_indices = [i for i, c in enumerate(col_stats) if c['is_numeric']]

            clabels = [
                str(data_rows[j][label_col]) if label_col is not None and label_col < len(data_rows[j]) else f"صف {j+1}"
                for j in range(min(50, total_rows))
            ]
            cdatasets = []
            for k, ci in enumerate(numeric_col_indices[:4]):
                raw = [data_rows[j][ci] if ci < len(data_rows[j]) else None for j in range(min(50, total_rows))]
                nums = [float(str(v).replace(',', '')) if _is_numeric(v) else None for v in raw]
                cdatasets.append({
                    'label': headers[ci],
                    'data': nums,
                    'color': palette[k % len(palette)],
                })

            sample = [[str(c) if c is not None else '' for c in r] for r in data_rows[:5]]
            result_sheets.append({
                'name': sheet_name,
                'total_rows': total_rows,
                'total_cols': len(headers),
                'headers': headers,
                'col_stats': col_stats,
                'sample': sample,
                'chart_labels': clabels,
                'chart_datasets': cdatasets,
                'has_charts': bool(cdatasets),
            })

        wb.close()
        return {'type': 'excel', 'sheets': result_sheets, 'sheet_count': len(sheets), 'error': None}
    except Exception as e:
        return {'type': 'excel', 'error': str(e)}


# ─── PDF ─────────────────────────────────────────────────────────────────────

def analyse_pdf(filepath):
    try:
        import pdfplumber
    except ImportError:
        return {'type': 'pdf', 'error': 'pip install pdfplumber'}

    try:
        with pdfplumber.open(filepath) as pdf:
            total_pages = len(pdf.pages)
            pages_data = []
            total_chars = 0
            all_tables = []

            # Speed optimized: only scan first 3 pages for strategic dashboard
            for page in pdf.pages[:3]:
                text = (page.extract_text() or '').strip()
                total_chars += len(text)
                pages_data.append({
                    'number': page.page_number,
                    'text': text[:300],
                    'has_tables': False, # Disabled for performance
                })

        # Try to extract numeric data from first table for chart
        chart_data = None
        # Tables extraction is too slow for real-time AI API. 
        # We rely on Groq Vision or text instead.

        return {
            'type': 'pdf',
            'total_pages': total_pages,
            'total_chars': total_chars,
            'pages_data': pages_data,
            'tables_found': len(all_tables),
            'chart_data': chart_data,
            'has_charts': chart_data is not None,
            'error': None,
        }
    except Exception as e:
        return {'type': 'pdf', 'error': str(e)}


# ─── IMAGE (NEW VISION ANALYSIS) ─────────────────────────────────────────────

def analyse_image(filepath):
    """Uses Groq Vision (Llama 3.2) to provide a strategic analysis of the image."""
    try:
        from groq import Groq
        import base64
        client = Groq(api_key=settings.GROQ_API_KEY)
        
        with open(filepath, "rb") as image_file:
            image_data = image_file.read()

        prompt = (
            "Analyze this image for an ERP strategic dashboard. "
            "Identify what the image is (invoice, site photo, chart, document). "
            "Provide a concise professional summary in Arabic. "
            "If it contains financial figures, list them clearly."
        )
        
        model = genai.GenerativeModel("gemini-2.5-flash")
        response = model.generate_content([
            {"mime_type": "image/jpeg" if "jpg" in filepath.lower() or "jpeg" in filepath.lower() else "image/png", "data": image_data},
            prompt
        ])
        description = response.text.strip()

        
        return {
            'type': 'image',
            'summary_ar': description,
            'has_charts': False,
            'error': None,
        }
    except Exception as e:
        return {'type': 'image', 'error': f"Multimodal Image Analysis Error: {str(e)}"}



# ─── Dispatcher ──────────────────────────────────────────────────────────────

def analyse_file(data_file_obj):
    """
    Dispatcher: Routes the file to the correct analyser based on extension.
    Optimized: Checks for cached analysis_result first.
    """
    if not data_file_obj:
        return {'type': 'unknown', 'error': 'ملف غير صالح'}
    
    # Bypass cached result to force real-time parsing from disk


    try:
        # 1. Attempt to get the absolute path
        if not data_file_obj or not hasattr(data_file_obj, 'file'):
             return {'type': 'unknown', 'error': 'ملف غير صالح'}
             
        filepath = data_file_obj.file.path
        
        # 2. Check if file physically exists on the server's disk
        if not os.path.exists(filepath):
            return {
                'type': 'unknown', 
                'error': f'المستند ({os.path.basename(filepath)}) غير موجود ماديًا على السيرفر. ربما تم حذفه أو لم يكتمل الرفع.'
            }
            
        ft = data_file_obj.file_type.lower()
        name = filepath.lower()
        ext = name.split('.')[-1]
        file_type = None

        if ext in ['pdf']: file_type = 'pdf'
        elif ext in ['xls', 'xlsx', 'xlsm']: file_type = 'excel'
        elif ext in ['csv']: file_type = 'csv'
        elif ext in ['jpg', 'jpeg', 'png', 'webp']: file_type = 'image'
        elif ext in ['txt', 'md', 'json', 'log', 'xml']: file_type = 'text'
        else: file_type = 'other'

        # 3. Dispatch to specific format analyser
        if file_type == 'csv':
            return analyse_csv(filepath)
        if file_type == 'excel':
            return analyse_excel(filepath)
        if file_type == 'pdf':
            return analyse_pdf(filepath)
        if file_type == 'image':
            return analyse_image(filepath)
        if file_type == 'text':
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read(3000) # Preview up to 3k chars
                return {
                    'type': 'text',
                    'pages_data': [{'number': 1, 'text': content}],
                    'error': None,
                    'has_charts': False
                }
            except Exception as e:
                return {'type': 'text', 'error': f'فشل قراءة النص: {str(e)}'}
            
        # Fallback for ALL other file types (e.g. Word, Zip, etc.)
        return {
            'type': 'other',
            'file_name': os.path.basename(filepath),
            'file_size_kb': round(os.path.getsize(filepath) / 1024, 2),
            'pages_data': [{'number': 1, 'text': f"مستند مرجعي: {os.path.basename(filepath)}"}],
            'error': None,
            'has_charts': False
        }

    except Exception as e:
        return {'type': 'unknown', 'error': f'خطأ تقني غير متوقع: {str(e)}'}
